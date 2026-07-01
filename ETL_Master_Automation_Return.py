
PIPELINE ORDER:
  STEP 1 → Flipkart PDF Processing  (FK_POD_Production)
  STEP 2 → FK Match                 (FKmatch)
  STEP 3 → Amazon Match             (Amazon_match)
  STEP 4 → Myntra Match             (Myntra_match)
  STEP 5 → Daily Summary Report     (Updated_Daily_Summary_Report)

HOW TO RUN:
  python ETL_Master_Automation.py

REQUIREMENTS:
  pip install pandas openpyxl xlsxwriter pdfplumber


from __future__ import annotations

# ================================================================
# IMPORTS
# ================================================================

import glob
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ── MIS FILE (Common for all portals) ──────────────────────────
MIS_FILE = r"C:\Users\Dilip\Desktop\ALL LOC MIS.xlsx"

# ── FLIPKART ───────────────────────────────────────────────────
FK_PDF_ROOT   = r"C:\Users\Dilip\Desktop\FK"
FK_MASTER_XLS = r"C:\Users\Dilip\Desktop\FK_POD_Master.xlsx"

# ── FLIPKART ORDER ID — RETURN AND SALES MIS ───────────────────
# ⚠️  ROZ SIRF YEH FILE NAME BADLO (date wala hissa)
# Sheet: "RETURN EN" | AN col = Reverse Tracking No | E col = Magento Order No
# Formula: =XLOOKUP(TrackingID, AN:AN, E:E,,0)
FK_ORDER_MIS  = r"C:\Users\Dilip\Downloads\RETURN AND SALES MIS 01-07-2026.xlsx"

# ── AMAZON ─────────────────────────────────────────────────────
AMAZON_INPUT  = r"C:\Users\Dilip\OneDrive - Happy Ecom Venture\Desktop\RDBNR\Amazon RDBNR.xlsx"

# ── MYNTRA ─────────────────────────────────────────────────────
# Agar file Downloads mein daily download hoti haitoh neeche wala use karo:
MYNTRA_INPUT  = r"C:\Users\Dilip\Downloads\Myntra 01-07-2026.xlsx"
# Ya agar same folder mein rehti hai:
# MYNTRA_INPUT  = r"C:\Users\Dilip\Desktop\Myntra.xlsx"

# ── REPORT DATES ───────────────────────────────────────────────
# Format: datetime(YEAR, MONTH, DAY)
REPORT_DATE_START = datetime(2026, 6, 30).date()
REPORT_DATE_END   = datetime(2026, 6, 30).date()

# ── OUTPUT FOLDER ──────────────────────────────────────────────
OUTPUT_FOLDER = Path.home() / "Downloads"

# ================================================================
# END OF SETTINGS — BAAKI KUCH MAT CHHUO
# ================================================================


# ── Auto-generated output file names ───────────────────────────
_ts = datetime.now().strftime("%d-%m-%Y_%H-%M")

FK_OUTPUT_FILE      = OUTPUT_FOLDER / f"FK_POD_Output_{_ts}.xlsx"
AMAZON_OUTPUT_FILE  = OUTPUT_FOLDER / f"Amazon_output_{_ts}.xlsx"
MYNTRA_OUTPUT_FILE  = OUTPUT_FOLDER / f"Myntra_output_{_ts}.xlsx"
SUMMARY_OUTPUT_FILE = OUTPUT_FOLDER / "Daily Summary Report.xlsx"

REPORT_GENERATED_ON = datetime.now().strftime("%d-%m-%Y %H:%M")


# ================================================================
# LOGGING SETUP
# ================================================================

Path("logs").mkdir(exist_ok=True)
logging.basicConfig(
    filename="logs/etl_master.log",
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)


# ================================================================
# STEP 0 — COMMON HELPERS
# ================================================================

def clean_text(x):
    """Whitespace + NBSP hataao, uppercase karo."""
    if pd.isna(x):
        return ""
    x = str(x)
    x = x.replace('\xa0', '')
    x = re.sub(r'\s+', '', x)
    return x.strip().upper()


def banner(text: str):
    line = "═" * 62
    print(f"\n{line}")
    print(f"  {text}")
    print(f"{line}")


def step_header(step: int, title: str):
    print(f"\n{'─'*62}")
    print(f"  STEP {step} ▶  {title}")
    print(f"{'─'*62}")


def build_mis_lookup(mis_file: str) -> dict:
    """
    MIS file load karo — column 6 (index 5) tracking IDs.
    Returns: {tracking_id: sheet_name}
    """
    print("  📂 MIS file load ho rahi hai...")
    all_sheets = pd.read_excel(mis_file, sheet_name=None, engine="openpyxl")

    combined_data = []
    for sheet_name, df in all_sheets.items():
        df.columns = [str(c).strip() for c in df.columns]
        if len(df.columns) <= 5:
            continue
        tracking_col = df.columns[5]
        temp = df[[tracking_col]].copy()
        temp.columns = ["Tracking_ID"]
        temp["Tracking_ID"] = temp["Tracking_ID"].map(clean_text)
        temp = temp[temp["Tracking_ID"] != ""]
        temp["Sheet_Name"] = sheet_name
        combined_data.append(temp)

    combined_df = pd.concat(combined_data, ignore_index=True)
    combined_df.drop_duplicates(subset=["Tracking_ID"], keep="first", inplace=True)

    lookup = dict(zip(combined_df["Tracking_ID"], combined_df["Sheet_Name"]))
    print(f"  ✅ MIS Ready | Tracking IDs: {len(lookup):,}")
    return lookup


# ================================================================
# STEP 1 — FLIPKART PDF PROCESSING
# ================================================================

DATA_SHEET      = "POD_Data"
PROCESSED_SHEET = "Processed_PDFs"
LOG_SHEET       = "Processing_Log"
TRACKING_REGEX  = re.compile(r"\bFM[A-Z0-9]{8,25}\b")


@dataclass
class RunStats:
    total_pdfs_found: int = 0
    new_pdfs_processed: int = 0
    modified_pdfs_reprocessed: int = 0
    skipped_unchanged: int = 0
    total_tracking_ids_extracted: int = 0
    total_records_added: int = 0
    duplicates_skipped: int = 0


def _ts_to_str(dt) -> str:
    if dt is None:
        return ""
    if hasattr(dt, "to_pydatetime"):
        dt = dt.to_pydatetime()
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _ensure_fk_workbook(master_file: str):
    p = Path(master_file)
    if not p.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = DATA_SHEET
        ws.append(["Date", "Warehouse", "Tracking_ID", "PDF_File", "Loaded_On"])
        ws2 = wb.create_sheet(PROCESSED_SHEET)
        ws2.append(["Warehouse", "PDF_File", "Last_Modified_Time", "Processed_On"])
        ws3 = wb.create_sheet(LOG_SHEET)
        ws3.append(["PDF_File", "Warehouse", "Processed_Date",
                    "Total_Tracking_Extracted", "Records_Loaded"])
        wb.save(master_file)
        return

    try:
        wb = load_workbook(master_file)
    except PermissionError:
        raise RuntimeError("FK_POD_Master.xlsx band karo aur dobara run karo.")

    required = {
        DATA_SHEET:      ["Date", "Warehouse", "Tracking_ID", "PDF_File", "Loaded_On"],
        PROCESSED_SHEET: ["Warehouse", "PDF_File", "Last_Modified_Time", "Processed_On"],
        LOG_SHEET:       ["PDF_File", "Warehouse", "Processed_Date",
                          "Total_Tracking_Extracted", "Records_Loaded"],
    }
    changed = False
    for sheet, cols in required.items():
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
            ws.append(cols)
            changed = True
    if changed:
        wb.save(master_file)


def _load_processed_index(master_file: str) -> dict:
    try:
        df = pd.read_excel(master_file, sheet_name=PROCESSED_SHEET)
    except Exception:
        return {}
    if df.empty:
        return {}
    idx = {}
    for _, r in df.iterrows():
        key = (str(r["Warehouse"]).strip(), str(r["PDF_File"]).strip())
        normalized_ts = _ts_to_str(r["Last_Modified_Time"])
        if key not in idx or normalized_ts > idx[key]:
            idx[key] = normalized_ts
    return idx


def _load_existing_tracking_keys(master_file: str) -> set:
    try:
        df = pd.read_excel(master_file, sheet_name=DATA_SHEET, dtype=str)
    except Exception:
        return set()
    if df.empty or not {"Warehouse", "Tracking_ID", "PDF_File"}.issubset(df.columns):
        return set()
    return set(zip(
        df["Warehouse"].fillna("").str.strip(),
        df["Tracking_ID"].fillna("").str.strip(),
        df["PDF_File"].fillna("").str.strip(),
    ))


def _append_df_sorted(rows: list, sheet_name: str, master_file: str):
    if not rows:
        return
    df = pd.DataFrame(rows)
    if sheet_name == DATA_SHEET:
        df[0] = pd.to_datetime(df[0], errors="coerce")
        df = df.sort_values(by=0, ascending=True, na_position="last")
    wb = load_workbook(master_file)
    ws = wb[sheet_name]
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(master_file)


def _update_processed_sheet(updates: dict, master_file: str):
    if not updates:
        return
    wb = load_workbook(master_file)
    ws = wb[PROCESSED_SHEET]
    existing_rows = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        wh  = str(row[0].value).strip() if row[0].value else ""
        pdf = str(row[1].value).strip() if row[1].value else ""
        existing_rows[(wh, pdf)] = row_num
    for (warehouse, pdf_name), (modified_dt, processed_on) in updates.items():
        key = (str(warehouse), str(pdf_name))
        if key in existing_rows:
            r = existing_rows[key]
            ws.cell(row=r, column=3).value = modified_dt
            ws.cell(row=r, column=4).value = processed_on
        else:
            ws.append([warehouse, pdf_name, modified_dt, processed_on])
    wb.save(master_file)


def _extract_date(text: str, pdf_name: str):
    m = re.search(r"Date:\s*([A-Za-z]{3}\s+\d{1,2},\s+\d{4})", text)
    if m:
        try:
            return pd.to_datetime(m.group(1)).date()
        except Exception:
            pass
    try:
        return datetime.strptime(Path(pdf_name).stem, "%d-%m-%Y").date()
    except Exception:
        return None


def _read_pdf(pdf_path: Path):
    tracking   = []
    text_parts = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                text_parts.append(txt)
                tracking.extend(TRACKING_REGEX.findall(txt))
    except Exception:
        logging.exception(f"Corrupt PDF: {pdf_path}")
        return [], ""
    return tracking, "\n".join(text_parts)


def step1_fk_pdf_processing() -> str:
    """STEP 1: FK PDFs → FK_POD_Master.xlsx mein data load karo."""
    step_header(1, "FLIPKART PDF PROCESSING")

    root = Path(FK_PDF_ROOT)
    if not root.exists():
        print(f"  ⚠️ FK folder nahi mila: {FK_PDF_ROOT} — Step 1 skip")
        return FK_MASTER_XLS

    _ensure_fk_workbook(FK_MASTER_XLS)
    processed_index        = _load_processed_index(FK_MASTER_XLS)
    existing_tracking_keys = _load_existing_tracking_keys(FK_MASTER_XLS)

    stats             = RunStats()
    pod_rows          = []
    processed_updates = {}
    log_rows          = []

    for wh_folder in sorted(p for p in root.iterdir() if p.is_dir()):
        warehouse = wh_folder.name
        for pdf_file in sorted(wh_folder.glob("*.pdf")):
            stats.total_pdfs_found += 1
            modified     = datetime.fromtimestamp(pdf_file.stat().st_mtime)
            modified_str = _ts_to_str(modified)
            key    = (warehouse, pdf_file.name)
            old_ts = processed_index.get(key)

            if old_ts and old_ts == modified_str:
                stats.skipped_unchanged += 1
                continue

            if old_ts:
                stats.modified_pdfs_reprocessed += 1
            else:
                stats.new_pdfs_processed += 1

            tracking_ids, text = _read_pdf(pdf_file)
            pod_date           = _extract_date(text, pdf_file.name)
            now                = datetime.now()

            new_count = 0
            dup_count = 0

            for tid in tracking_ids:
                combo = (str(warehouse), str(tid), str(pdf_file.name))
                if combo in existing_tracking_keys:
                    dup_count += 1
                    continue
                pod_rows.append([pod_date, warehouse, tid, pdf_file.name, now])
                existing_tracking_keys.add(combo)
                new_count += 1

            processed_updates[key] = (modified, now)
            log_rows.append([pdf_file.name, warehouse, now, len(tracking_ids), new_count])

            stats.total_tracking_ids_extracted += len(tracking_ids)
            stats.total_records_added          += new_count
            stats.duplicates_skipped           += dup_count

    _append_df_sorted(pod_rows, DATA_SHEET, FK_MASTER_XLS)
    _update_processed_sheet(processed_updates, FK_MASTER_XLS)
    _append_df_sorted(log_rows, LOG_SHEET, FK_MASTER_XLS)

    print(f"  📄 Total PDFs Found          : {stats.total_pdfs_found}")
    print(f"  ⏭  Unchanged (Skipped)       : {stats.skipped_unchanged}")
    print(f"  🆕 New PDFs Processed        : {stats.new_pdfs_processed}")
    print(f"  🔄 Modified PDFs Reprocessed : {stats.modified_pdfs_reprocessed}")
    print(f"  🔑 Tracking IDs Extracted    : {stats.total_tracking_ids_extracted}")
    print(f"  ➕ Records Added             : {stats.total_records_added}")
    print(f"  🔁 Duplicates Skipped        : {stats.duplicates_skipped}")
    print(f"  ✅ FK Master Updated: {FK_MASTER_XLS}")
    return FK_MASTER_XLS


# ================================================================
# STEP 2 — FK ORDER ID LOOKUP HELPER
# ================================================================

def build_fk_order_lookup() -> dict:
    """
    RETURN AND SALES MIS file se FK Order ID lookup dict banao.

    Excel formula:
    =XLOOKUP(TrackingID, AN:AN, E:E,,0)
    """

    if not os.path.exists(FK_ORDER_MIS):
        print(f"  ⚠️ FK Order MIS file nahi mili: {FK_ORDER_MIS}")
        print("     Order_ID column 'Not Found' rahega.")
        return {}

    print(f"  📂 FK Order MIS load ho rahi hai: {Path(FK_ORDER_MIS).name}")

    try:
        df = pd.read_excel(
            FK_ORDER_MIS,
            sheet_name="RETURN EN",
            engine="openpyxl",
            dtype=str
        )
    except Exception as e:
        print(f"  ⚠️ FK Order MIS load error: {e}")
        return {}

    if df.shape[1] < 40:
        print(f"  ⚠️ RETURN EN sheet mein expected 40+ columns nahi mile (Found {df.shape[1]})")
        return {}

    an_col = df.iloc[:, 39].fillna("")   # Reverse Tracking No
    n_col  = df.iloc[:, 13].fillna("")   # Tracking No (fallback)
    e_col  = df.iloc[:, 4].fillna("")    # Magento Order No

    lookup = {}

    for an_val, n_val, e_val in zip(an_col, n_col, e_col):

        order_id = str(e_val).strip()

        if not order_id or order_id.upper() == "NAN":
            continue

        # Primary Lookup (AN)
        an_clean = clean_text(an_val)
        if an_clean and an_clean not in lookup:
            lookup[an_clean] = order_id

        # Secondary Lookup (N)
        n_clean = clean_text(n_val)
        if n_clean and n_clean not in lookup:
            lookup[n_clean] = order_id

    print(f"  ✅ FK Order Lookup Ready | Entries: {len(lookup):,}")

    return lookup
# ================================================================
# STEP 2 — FK MATCH (with Order ID)
# ================================================================

def step2_fk_match(lookup_dict: dict) -> Path:
    """
    STEP 2: FK_POD_Master.xlsx ka POD_Data → MIS match + Order_ID lookup.

    Output columns:
      Found_In_Sheet      = MIS sheet name (ya 'Not Found')
      Matched_Tracking_ID = matched tracking ID (ya 'Not Found')
      Order_ID            = OI:XXXXXXXXXXXXXXXXXX (ya 'Not Found')
    """
    step_header(2, "FLIPKART MATCH (MIS + Order ID)")

    try:
        input_df = pd.read_excel(FK_MASTER_XLS, sheet_name="POD_Data", engine="openpyxl")
    except ValueError:
        print(f"  ⚠️ 'POD_Data' sheet nahi mili FK Master mein — Step 2 skip")
        return FK_OUTPUT_FILE
    except Exception as e:
        print(f"  ⚠️ FK Master load error: {e} — Step 2 skip")
        return FK_OUTPUT_FILE

    input_df.columns = [str(c).strip() for c in input_df.columns]

    if "Tracking_ID" not in input_df.columns:
        print("  ⚠️ 'Tracking_ID' column nahi mila — Step 2 skip")
        return FK_OUTPUT_FILE

    input_df["Tracking_ID"] = input_df["Tracking_ID"].map(clean_text)

    # ── MIS Sheet Match ────────────────────────────────────────
    match_result = input_df["Tracking_ID"].map(lookup_dict)
    input_df["Found_In_Sheet"]      = match_result.fillna("Not Found")
    input_df["Matched_Tracking_ID"] = (
        input_df["Tracking_ID"].where(match_result.notna()).fillna("Not Found")
    )
    found     = match_result.notna().sum()
    not_found = match_result.isna().sum()
    print(f"  ✅ MIS Matched    : {found:,}")
    print(f"  ❌ MIS Not Found  : {not_found:,}")

    # ── Order ID Lookup ────────────────────────────────────────
    fk_order_lookup = build_fk_order_lookup()
    input_df["Order_ID"] = (
        input_df["Tracking_ID"].map(fk_order_lookup).fillna("Not Found")
    )
    order_found     = input_df["Order_ID"].ne("Not Found").sum()
    order_not_found = input_df["Order_ID"].eq("Not Found").sum()
    print(f"  🔑 Order ID Found : {order_found:,}")
    print(f"  ❌ Order ID Miss  : {order_not_found:,}")

    with pd.ExcelWriter(FK_OUTPUT_FILE, engine="openpyxl") as writer:
        input_df.to_excel(writer, sheet_name="POD_Data", index=False)

    print(f"  💾 FK Output: {FK_OUTPUT_FILE}")
    return FK_OUTPUT_FILE


# ================================================================
# STEP 3 — AMAZON MATCH
# ================================================================

def step3_amazon_match(lookup_dict: dict) -> Path:
    """STEP 3: Amazon RDBNR.xlsx ke 2 sheets → MIS se match karo."""
    step_header(3, "AMAZON MATCH (MIS Lookup)")

    if not os.path.exists(AMAZON_INPUT):
        print(f"  ⚠️ Amazon file nahi mili: {AMAZON_INPUT} — Step 3 skip")
        return AMAZON_OUTPUT_FILE

    sheets = ["Return initiate mail", "Succesfull delivered"]

    with pd.ExcelWriter(AMAZON_OUTPUT_FILE, engine="xlsxwriter") as writer:
        for sheet in sheets:
            print(f"  ⚡ Processing: {sheet}")
            try:
                df = pd.read_excel(AMAZON_INPUT, sheet_name=sheet)
            except Exception as e:
                print(f"  ⚠️ Sheet '{sheet}' skip: {e}")
                continue

            df.columns = [str(c).strip() for c in df.columns]
            col_D = df.columns[3]  # Column D = index 3
            df[col_D] = df[col_D].map(clean_text)

            match_D = df[col_D].map(lookup_dict)
            df["Found_In_Sheet"]      = match_D.fillna("Not Found")
            df["Matched_Tracking_ID"] = df[col_D].where(match_D.notna()).fillna("Not Found")

            found     = match_D.notna().sum()
            not_found = match_D.isna().sum()
            print(f"     ✅ Matched: {found:,}  |  ❌ Not Found: {not_found:,}")

            df.to_excel(writer, sheet_name=sheet, index=False)

    print(f"  💾 Amazon Output: {AMAZON_OUTPUT_FILE}")
    return AMAZON_OUTPUT_FILE


# ================================================================
# STEP 4 — MYNTRA MATCH
# ================================================================

def step4_myntra_match(lookup_dict: dict) -> Path:
    """STEP 4: Myntra file ke columns W, X, Y → MIS se match karo."""
    step_header(4, "MYNTRA MATCH (MIS Lookup)")

    if not os.path.exists(MYNTRA_INPUT):
        print(f"  ⚠️ Myntra file nahi mili: {MYNTRA_INPUT} — Step 4 skip")
        return MYNTRA_OUTPUT_FILE

    df = pd.read_excel(MYNTRA_INPUT)
    df.columns = [str(c).strip() for c in df.columns]

    # Columns W (22), X (23), Y (24)
    col_W = df.columns[22]
    col_X = df.columns[23]
    col_Y = df.columns[24]

    df[col_W] = df[col_W].map(clean_text)
    df[col_X] = df[col_X].map(clean_text)
    df[col_Y] = df[col_Y].map(clean_text)

    match_W = df[col_W].map(lookup_dict)
    match_X = df[col_X].map(lookup_dict)
    match_Y = df[col_Y].map(lookup_dict)

    # First match wins (W → X → Y)
    df["Found_In_Sheet"] = match_W.fillna(match_X).fillna(match_Y).fillna("Not Found")
    df["Matched_Tracking_ID"] = (
        df[col_W].where(match_W.notna())
        .fillna(df[col_X].where(match_X.notna()))
        .fillna(df[col_Y].where(match_Y.notna()))
        .fillna("Not Found")
    )

    found     = df["Found_In_Sheet"].ne("Not Found").sum()
    not_found = df["Found_In_Sheet"].eq("Not Found").sum()
    print(f"  ✅ Matched   : {found:,}")
    print(f"  ❌ Not Found : {not_found:,}")

    df.to_excel(MYNTRA_OUTPUT_FILE, index=False)
    print(f"  💾 Myntra Output: {MYNTRA_OUTPUT_FILE}")
    return MYNTRA_OUTPUT_FILE


# ================================================================
# STEP 5 — DAILY SUMMARY REPORT
# ================================================================

# ── Styling constants ───────────────────────────────────────────
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
HEADER_BLUE  = "1F4E78"
SUBHDR_BLUE  = "D9EAF7"
GRAND_GREEN  = "E2EFDA"
FK_ORANGE    = "F4B942"
WHITE        = "FFFFFF"


def _apply_title_style(cell, text, color=HEADER_BLUE):
    cell.value = text
    cell.font  = Font(size=14, bold=True, color=WHITE)
    cell.fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_header_style(cell, color=SUBHDR_BLUE):
    cell.font  = Font(bold=True)
    cell.fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin


def _write_section(ws, df, start_row, title_text, title_color=HEADER_BLUE, subhdr_color=SUBHDR_BLUE):
    if df.empty:
        return start_row
    cols     = list(df.columns)
    num_cols = len(cols)
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=num_cols)
    _apply_title_style(ws.cell(start_row, 1), title_text, color=title_color)
    for col_num, col_name in enumerate(cols, 1):
        cell = ws.cell(start_row + 1, col_num)
        cell.value = col_name
        _apply_header_style(cell, color=subhdr_color)
    for row_offset, row in enumerate(df.values, start=2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(start_row + row_offset, col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
            if str(row[0]).lower() == "grand total":
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=GRAND_GREEN, end_color=GRAND_GREEN, fill_type="solid")
    return start_row + 2 + len(df)


def _empty_summary():
    return pd.DataFrame(columns=["Portal", "Location", "Date", "Report Date",
                                  "Today Initiated", "PH Received", "Diff"])


def _process_amazon_summary(amazon_file: Path):
    try:
        df = pd.read_excel(amazon_file, sheet_name="Succesfull delivered")
        required = ["Date", "Location", "Found_In_Sheet"]
        missing  = [c for c in required if c not in df.columns]
        if missing:
            print(f"  ⚠️ Amazon summary skip. Missing: {missing}")
            return _empty_summary(), pd.DataFrame()

        df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce").dt.date
        df = df[(df["Date"] >= REPORT_DATE_START) & (df["Date"] <= REPORT_DATE_END)]
        if df.empty:
            print("  ⚠️ Amazon: date range mein koi record nahi")
            return _empty_summary(), pd.DataFrame()

        df["Location"]       = df["Location"].astype(str).str.strip().str.upper()
        df["Found_In_Sheet"] = df["Found_In_Sheet"].astype(str).str.strip().str.lower()

        mismatch_df = df[df["Found_In_Sheet"] == "not found"].copy()
        if not mismatch_df.empty:
            mismatch_df.insert(0, "Portal", "Amazon")
            mismatch_df.insert(1, "Report Generated On", REPORT_GENERATED_ON)

        summary_df = (
            df.groupby(["Location", "Date"])
            .agg(
                Today_Initiated=("Location", "size"),
                Diff=("Found_In_Sheet", lambda x: (x == "not found").sum())
            )
            .reset_index()
        )
        summary_df["PH Received"] = summary_df["Today_Initiated"] - summary_df["Diff"]
        summary_df["Portal"]      = "Amazon"
        summary_df["Report Date"] = f"{REPORT_DATE_START} to {REPORT_DATE_END}"
        summary_df["Date"]        = pd.to_datetime(summary_df["Date"]).dt.strftime("%d-%m-%Y")
        summary_df.rename(columns={"Today_Initiated": "Today Initiated"}, inplace=True)
        return summary_df[["Portal", "Location", "Date", "Report Date",
                            "Today Initiated", "PH Received", "Diff"]], mismatch_df

    except Exception as e:
        print(f"  ⚠️ Amazon summary error: {e}")
        return _empty_summary(), pd.DataFrame()


def _process_myntra_summary(myntra_file: Path):
    try:
        df = pd.read_excel(myntra_file)
        required = ["deliver_to_seller_date", "warehouse_id", "Found_In_Sheet"]
        missing  = [c for c in required if c not in df.columns]
        if missing:
            print(f"  ⚠️ Myntra summary skip. Missing: {missing}")
            return _empty_summary(), pd.DataFrame()

        df["deliver_to_seller_date"] = pd.to_datetime(
            df["deliver_to_seller_date"], dayfirst=True, errors="coerce"
        ).dt.date
        df = df[(df["deliver_to_seller_date"] >= REPORT_DATE_START) &
                (df["deliver_to_seller_date"] <= REPORT_DATE_END)]
        if df.empty:
            print("  ⚠️ Myntra: date range mein koi record nahi")
            return _empty_summary(), pd.DataFrame()

        df["warehouse_id"]   = df["warehouse_id"].astype(str).str.strip().str.upper()
        df["Found_In_Sheet"] = df["Found_In_Sheet"].astype(str).str.strip().str.lower()

        mismatch_df = df[df["Found_In_Sheet"] == "not found"].copy()
        if not mismatch_df.empty:
            mismatch_df.insert(0, "Portal", "Myntra")
            mismatch_df.insert(1, "Report Generated On", REPORT_GENERATED_ON)

        summary_df = (
            df.groupby(["warehouse_id", "deliver_to_seller_date"])
            .agg(
                Today_Initiated=("warehouse_id", "size"),
                Diff=("Found_In_Sheet", lambda x: (x == "not found").sum())
            )
            .reset_index()
        )
        summary_df["PH Received"] = summary_df["Today_Initiated"] - summary_df["Diff"]
        summary_df["Portal"]      = "Myntra"
        summary_df["Report Date"] = f"{REPORT_DATE_START} to {REPORT_DATE_END}"
        summary_df["deliver_to_seller_date"] = pd.to_datetime(
            summary_df["deliver_to_seller_date"]
        ).dt.strftime("%d-%m-%Y")
        summary_df.rename(columns={
            "warehouse_id": "Location",
            "deliver_to_seller_date": "Date",
            "Today_Initiated": "Today Initiated"
        }, inplace=True)
        return summary_df[["Portal", "Location", "Date", "Report Date",
                            "Today Initiated", "PH Received", "Diff"]], mismatch_df

    except Exception as e:
        print(f"  ⚠️ Myntra summary error: {e}")
        return _empty_summary(), pd.DataFrame()


def _process_flipkart_summary(fk_file: Path):
    try:
        df = pd.read_excel(fk_file, sheet_name="POD_Data")
        df.columns = [str(c).strip() for c in df.columns]

        col_date     = "Date"
        col_location = "Warehouse"
        col_found    = "Found_In_Sheet"

        required = [col_date, col_location, col_found]
        missing  = [c for c in required if c not in df.columns]
        if missing:
            print(f"  ⚠️ Flipkart summary skip. Missing: {missing}")
            return _empty_summary(), pd.DataFrame()

        df[col_date] = pd.to_datetime(df[col_date], dayfirst=True, errors="coerce").dt.date
        df = df[(df[col_date] >= REPORT_DATE_START) & (df[col_date] <= REPORT_DATE_END)]
        if df.empty:
            print("  ⚠️ Flipkart: date range mein koi record nahi")
            return _empty_summary(), pd.DataFrame()

        df[col_location] = df[col_location].astype(str).str.strip().str.upper()
        df[col_found]    = df[col_found].astype(str).str.strip().str.lower()

        # ── Mismatch: Order_ID column bhi include karo ─────────
        mismatch_df = df[df[col_found] == "not found"].copy()
        if not mismatch_df.empty:
            mismatch_df.insert(0, "Portal", "Flipkart")
            mismatch_df.insert(1, "Report Generated On", REPORT_GENERATED_ON)
            # Order_ID column ensure karo (agar step2 ne add kiya toh aayega)
            if "Order_ID" not in mismatch_df.columns:
                mismatch_df["Order_ID"] = "Not Found"

        summary_df = (
            df.groupby([col_location, col_date])
            .agg(
                Today_Initiated=(col_location, "size"),
                Diff=(col_found, lambda x: (x == "not found").sum())
            )
            .reset_index()
        )
        summary_df["PH Received"] = summary_df["Today_Initiated"] - summary_df["Diff"]
        summary_df["Portal"]      = "Flipkart"
        summary_df["Report Date"] = f"{REPORT_DATE_START} to {REPORT_DATE_END}"
        summary_df[col_date]      = pd.to_datetime(summary_df[col_date]).dt.strftime("%d-%m-%Y")
        summary_df.rename(columns={
            col_location: "Location",
            col_date: "Date",
            "Today_Initiated": "Today Initiated"
        }, inplace=True)
        return summary_df[["Portal", "Location", "Date", "Report Date",
                            "Today Initiated", "PH Received", "Diff"]], mismatch_df

    except Exception as e:
        print(f"  ⚠️ Flipkart summary error: {e}")
        return _empty_summary(), pd.DataFrame()


def step5_daily_summary_report(
    fk_output: Path, amazon_output: Path, myntra_output: Path
):
    """STEP 5: Sab portal outputs → Ek Daily Summary Report banao."""
    step_header(5, "DAILY SUMMARY REPORT")

    # Process each portal
    amazon_df,   amazon_mismatch   = _process_amazon_summary(amazon_output)
    myntra_df,   myntra_mismatch   = _process_myntra_summary(myntra_output)
    flipkart_df, flipkart_mismatch = _process_flipkart_summary(fk_output)

    # Combine summary
    all_summaries = [amazon_df, myntra_df, flipkart_df]
    summary_df = pd.concat(all_summaries, ignore_index=True)

    if not summary_df.empty:
        grand_total = {
            "Portal": "Grand Total", "Location": "", "Date": "", "Report Date": "",
            "Today Initiated": summary_df["Today Initiated"].sum(),
            "PH Received":     summary_df["PH Received"].sum(),
            "Diff":            summary_df["Diff"].sum()
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)

    # Workbook create/load
    if os.path.exists(SUMMARY_OUTPUT_FILE):
        wb = load_workbook(SUMMARY_OUTPUT_FILE)
    else:
        wb = Workbook()
        if wb.active:
            wb.active.title = "Summary Report"

    # ── Summary Sheet ──
    if "Summary Report" in wb.sheetnames:
        ws_sum = wb["Summary Report"]
        sum_start = ws_sum.max_row + 3
    else:
        ws_sum = wb.create_sheet("Summary Report", 0)
        sum_start = 1

    _write_section(ws_sum, summary_df, sum_start,
                   f"Summary Report Generated On : {REPORT_GENERATED_ON}")

    for col, width in zip("ABCDEFG", [18, 18, 15, 30, 20, 15, 10]):
        ws_sum.column_dimensions[col].width = width

    # ── Amazon Mismatch ──
    if not amazon_mismatch.empty:
        ws = wb["Amazon Mismatch"] if "Amazon Mismatch" in wb.sheetnames else wb.create_sheet("Amazon Mismatch")
        start = ws.max_row + 3 if ws.max_row > 1 else 1
        _write_section(ws, amazon_mismatch, start,
                       f"Amazon Mismatch — Report Generated On : {REPORT_GENERATED_ON}")
        print(f"  ✅ Amazon Mismatch: {len(amazon_mismatch)} rows")

    # ── Myntra Mismatch ──
    if not myntra_mismatch.empty:
        ws = wb["Myntra Mismatch"] if "Myntra Mismatch" in wb.sheetnames else wb.create_sheet("Myntra Mismatch")
        start = ws.max_row + 3 if ws.max_row > 1 else 1
        _write_section(ws, myntra_mismatch, start,
                       f"Myntra Mismatch — Report Generated On : {REPORT_GENERATED_ON}")
        print(f"  ✅ Myntra Mismatch: {len(myntra_mismatch)} rows")

    # ── Flipkart Mismatch ──
    if not flipkart_mismatch.empty:
        ws = wb["FK Mismatch"] if "FK Mismatch" in wb.sheetnames else wb.create_sheet("FK Mismatch")
        start = ws.max_row + 3 if ws.max_row > 1 else 1
        _write_section(ws, flipkart_mismatch, start,
                       f"Flipkart Mismatch — Report Generated On : {REPORT_GENERATED_ON}",
                       title_color=FK_ORANGE)
        print(f"  ✅ FK Mismatch: {len(flipkart_mismatch)} rows")

    wb.save(SUMMARY_OUTPUT_FILE)
    print(f"  💾 Summary Report: {SUMMARY_OUTPUT_FILE}")


# ================================================================
# MAIN — PIPELINE ORCHESTRATOR
# ================================================================

def main():
    start_time = datetime.now()

    banner("HAPPY ECOM — MASTER ETL AUTOMATION v1.0")
    print(f"  🕐 Started At : {start_time.strftime('%d-%m-%Y %H:%M:%S')}")
    print(f"  📅 Report Date: {REPORT_DATE_START} to {REPORT_DATE_END}")

    # ── STEP 1: FK PDF Processing ──────────────────────────────
    step1_fk_pdf_processing()

    # ── MIS Lookup (SINGLE LOAD — used in Steps 2, 3, 4) ──────
    print(f"\n  📋 Loading MIS (common for all portals)...")
    lookup_dict = build_mis_lookup(MIS_FILE)

    # ── STEP 2: FK Match ───────────────────────────────────────
    fk_output = step2_fk_match(lookup_dict)

    # ── STEP 3: Amazon Match ───────────────────────────────────
    amazon_output = step3_amazon_match(lookup_dict)

    # ── STEP 4: Myntra Match ───────────────────────────────────
    myntra_output = step4_myntra_match(lookup_dict)

    # ── STEP 5: Daily Summary Report ───────────────────────────
    step5_daily_summary_report(fk_output, amazon_output, myntra_output)

    # ── FINAL SUMMARY ──────────────────────────────────────────
    end_time = datetime.now()
    elapsed  = (end_time - start_time).seconds

    banner("✅ ETL COMPLETE")
    print(f"  🕐 Finished At : {end_time.strftime('%d-%m-%Y %H:%M:%S')}")
    print(f"  ⏱  Total Time  : {elapsed} seconds")
    print()
    print(f"  📄 FK Output      → {FK_OUTPUT_FILE}")
    print(f"  📄 Amazon Output  → {AMAZON_OUTPUT_FILE}")
    print(f"  📄 Myntra Output  → {MYNTRA_OUTPUT_FILE}")
    print(f"  📄 Summary Report → {SUMMARY_OUTPUT_FILE}")
    print()


if __name__ == "__main__":
    main()
